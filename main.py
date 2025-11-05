from sqlalchemy import create_engine, inspect, text
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

# Load .env file
load_dotenv()  # By default it looks for a .env file in the current directory
DB_URI = os.getenv("DB_URI")
engine = create_engine(DB_URI)
insp = inspect(engine)

# --- schema and tables to export ---
schema = "public"
table_names = ["companies","attendances","admins","companies_menus","admin_use_cases","companies_objects","companies_menus_objects","companies_roles","devices","companies_role_menus","holiday_calendar","locations","log_palm_record","malut_attendance","master_pob","master_position_type","master_unit_organization","master_position","master_grade","notifications_admins","menus_objects","notif_messages","role_menus","notifications","menus","objects","roles","shift_rules","shift_details","shift_rule_holidays","shift_schedule","schema_migrations","user_grade","sofifi_employees","user_position","shifts","users","use_cases","devices_use_cases","config","master_group_pppk","location_unor","shift_rule_unit","log_login","user_admin","user_palm","role_user_admin","master_group_rank","user_group_rank","companies_location","log_file_import","companies_shift","companies_shift_detail"]

# --- reusable SQL queries ---
pk_q = text("""
SELECT kcu.column_name
FROM information_schema.table_constraints tc
JOIN information_schema.key_column_usage kcu
  ON tc.constraint_name = kcu.constraint_name
  AND tc.table_schema = kcu.table_schema
WHERE tc.constraint_type = 'PRIMARY KEY'
  AND tc.table_schema = :schema
  AND tc.table_name = :table;
""")

fk_q = text("""
SELECT
  kcu.column_name,
  ccu.table_schema AS foreign_table_schema,
  ccu.table_name   AS foreign_table_name,
  ccu.column_name  AS foreign_column_name,
  tc.constraint_name
FROM information_schema.table_constraints tc
JOIN information_schema.key_column_usage kcu
  ON tc.constraint_name = kcu.constraint_name
  AND tc.table_schema = kcu.table_schema
JOIN information_schema.constraint_column_usage ccu
  ON tc.constraint_name = ccu.constraint_name
  AND tc.table_schema = ccu.constraint_schema
WHERE tc.constraint_type = 'FOREIGN KEY'
  AND tc.table_schema = :schema
  AND tc.table_name = :table;
""")

col_meta_q = text("""
SELECT
  column_name,
  col_description((table_schema||'.'||table_name)::regclass::oid, ordinal_position) AS column_comment,
  character_maximum_length
FROM information_schema.columns
WHERE table_schema = :schema AND table_name = :table;
""")

# --- begin export process ---
excel_path = "db_table_docs.xlsx"
writer = pd.ExcelWriter(excel_path, engine="openpyxl")

for table_name in table_names:
    print(f"📄 Processing {table_name}...")

    columns = insp.get_columns(table_name, schema=schema)

    with engine.connect() as conn:
        pk_rows = conn.execute(pk_q, {"schema": schema, "table": table_name}).fetchall()
        fk_rows = conn.execute(fk_q, {"schema": schema, "table": table_name}).fetchall()
        meta_rows = conn.execute(col_meta_q, {"schema": schema, "table": table_name}).fetchall()

    pk_set = {r[0] for r in pk_rows}
    fk_map = defaultdict(list)
    for col_name, foreign_schema, foreign_table, foreign_col, constraint_name in fk_rows:
        fk_map[col_name].append(f"{foreign_schema}.{foreign_table}({foreign_col})[{constraint_name}]")

    comment_map = {}
    length_map = {}
    for column_name, column_comment, char_max_len in meta_rows:
        comment_map[column_name] = column_comment or ""
        length_map[column_name] = int(char_max_len) if char_max_len is not None else ""

    rows = []
    for i, col in enumerate(columns, start=1):
        col_name = col["name"]
        pk_flag = "PK" if col_name in pk_set else ""
        fk_flag = "FK" if col_name in fk_map else ""
        pkfk = ", ".join(filter(None, [pk_flag, fk_flag]))

        rows.append({
            "No": i,
            "Column": col_name,
            "Desc": comment_map.get(col_name, ""),
            "Null?": col["nullable"],
            "Data Type": str(col["type"]),
            "Length": length_map.get(col_name, ""),
            "PK / FK": pkfk,
        })

    df = pd.DataFrame(rows)

    # Write basic DataFrame
    df.to_excel(writer, sheet_name=table_name[:31], index=False)

writer.close()

# Define a thin border style for the table grid
thin = Side(border_style="thin", color="000000")
border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

wb = load_workbook(excel_path)

for table_name in table_names:
    ws = wb[table_name[:31]]

    # Insert two rows at top for "Table Name" and "Description"
    ws.insert_rows(1, amount=2)

    # --- Table Name section ---
    ws["A1"] = "Table Name"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Merge B1 to F1 for editable table name cell
    ws.merge_cells("B1:G1")
    ws["B1"] = table_name
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Description section ---
    ws["A2"] = "Description"
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Merge B2 to F2 for editable description area
    ws.merge_cells("B2:G2")
    ws["B2"] = ""  # leave empty
    ws["B2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # --- Header row styling ---
    header_row_idx = 3
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    # --- Apply border to entire data area ---
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_all
            cell.alignment = Alignment(vertical="center")

    # --- Adjust column widths ---
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[col_letter].width = min(max(10, max_length + 2), 60)

    # --- Apply border around Table Name / Description blocks ---
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = border_all

wb.save(excel_path)
print(f"✅ Exported, merged, and bordered: {excel_path}")